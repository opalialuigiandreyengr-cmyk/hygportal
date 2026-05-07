import re
import csv
import io
from datetime import date, datetime

from flask import Blueprint, flash, redirect, render_template, request, session, url_for, Response, jsonify
from flask_login import current_user
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from flask_login import login_required
from werkzeug.security import generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import db
from .helpers import (
    create_notification,
    _parse_date,
    _parse_int,
    _save_company_logo,
    _save_employee_photo,
    roles_required,
    sync_department_name,
)
from .models import Company, Department, Employee, User, EsarfApprover, EsarfRequest, LeaveRequest, LeaveApprover, DiscountRequest, ProductChargeRequest, PerkApprover

admin = Blueprint("admin", __name__)

ESARF_STATUS_PENDING = "Pending"
ESARF_STATUS_DEPT_MGR_APPROVED = "Dept Mgr Approved"
ESARF_STATUS_DEPT_MGR_OPS_APPROVED = "Dept Mgr Ops Approved"
ESARF_STATUS_APPROVED = "Approved"
ESARF_STATUS_REJECTED = "Rejected"

ESARF_APPROVER_ROLES = [
    ("dept manager", "Department Manager"),
    ("operation", "Operations Manager"),
    ("general manager", "General Manager"),
]
ESARF_APPROVER_ROLE_KEYS = {role for role, _label in ESARF_APPROVER_ROLES}

LEAVE_STATUS_PENDING = "Pending"
LEAVE_STATUS_DEPT_HR_APPROVED = "Dept/HR Approved"
LEAVE_STATUS_DEPT_HR_OPS_APPROVED = "Dept/HR Ops Approved"
LEAVE_STATUS_APPROVED = "Approved"
LEAVE_STATUS_REJECTED = "Rejected"

LEAVE_APPROVER_ROLES = [
    ("department", "Department Approver"),
    ("hr", "HR Approver"),
    ("operation", "Operations Approver"),
    ("general manager", "General Manager Approver"),
]
LEAVE_APPROVER_ROLE_KEYS = {role for role, _label in LEAVE_APPROVER_ROLES}

GM_REQUIRED_POSITIONS = {
    "it manager",
    "accounting head",
    "cluster manager",
    "hr head",
    "logistics manager",
    "maintenance manager",
    "payroll officer",
    "inventory manager",
    "general manager",
    "operation manager",
    "area manager",
    "training officer",
    "marketing manager",
    "internal auditor",
    "finance manager",
    "purchasing officer",
    "warehouse supervisor",
    "branch manager",
    "team leader",
    "supervisor",
}


def _requires_general_manager_approval(submitter_user):
    employee = submitter_user.employee if submitter_user else None
    position = (employee.position or "").strip().lower() if employee else ""
    if not position:
        return False
    return position in GM_REQUIRED_POSITIONS


def _leave_total_days(leave_request):
    if not leave_request or not leave_request.start_date or not leave_request.end_date:
        return 0
    return max(0, (leave_request.end_date - leave_request.start_date).days + 1)


def _deductible_leave_days(leave_request):
    leave_type_raw = (leave_request.leave_type or "").strip()
    leave_type = leave_type_raw.lower()
    total_days = _leave_total_days(leave_request)
    if total_days <= 0:
        return 0
    if leave_type.startswith("without pay"):
        return 0
    if leave_type.startswith("with pay"):
        return total_days
    if leave_type.startswith("both"):
        match = re.search(r"with\s*pay\s*:\s*(\d+)", leave_type_raw, re.IGNORECASE)
        if match:
            return max(0, int(match.group(1)))
        return total_days
    return total_days


def _apply_leave_credit_deduction(leave_request):
    submitter = leave_request.submitted_by_user if leave_request else None
    if not submitter:
        return 0
    if submitter.leave_credits is None:
        submitter.leave_credits = 0
    deducted_days = _deductible_leave_days(leave_request)
    if deducted_days <= 0:
        return 0
    submitter.leave_credits = max(0, int(submitter.leave_credits) - deducted_days)
    return deducted_days


def _normalize_department_name(department_name):
    return " ".join((department_name or "").strip().lower().split())


def _user_department(user):
    if not user or not user.employee:
        return ""
    return _normalize_department_name(user.employee.department)


def _esarf_submitter_department(esarf_request):
    submitter = esarf_request.submitted_by_user if esarf_request else None
    if not submitter or not submitter.employee:
        return ""
    return _normalize_department_name(submitter.employee.department)


def _current_esarf_approver_assignment():
    return EsarfApprover.query.filter_by(user_id=current_user.id).first()


def _current_esarf_workflow_role():
    user_role = (current_user.role or "").strip().lower()
    if user_role in {"admin", "timekeeper"}:
        return user_role

    assignment = _current_esarf_approver_assignment()
    if assignment and assignment.approver_role in ESARF_APPROVER_ROLE_KEYS:
        return assignment.approver_role

    return ""


def _current_user_can_manage_esarf():
    return _current_esarf_workflow_role() in {
        "admin",
        "timekeeper",
        *ESARF_APPROVER_ROLE_KEYS,
    }


def _department_manager_department():
    assignment = _current_esarf_approver_assignment()
    if assignment and assignment.approver_role == "dept manager":
        return _normalize_department_name(assignment.department_name)
    return _user_department(current_user)


def _department_manager_can_access_esarf(esarf_request):
    manager_department = _department_manager_department()
    request_department = _esarf_submitter_department(esarf_request)
    return bool(manager_department and request_department and manager_department == request_department)


def _scope_esarf_query_for_current_user(esarf_request_query):
    current_role = _current_esarf_workflow_role()
    if current_role != "dept manager":
        return esarf_request_query

    manager_department = _department_manager_department()
    if not manager_department:
        return esarf_request_query.filter(EsarfRequest.id.is_(None))

    return (
        esarf_request_query
        .join(EsarfRequest.submitted_by_user)
        .join(User.employee)
        .filter(func.lower(func.trim(Employee.department)) == manager_department)
    )


def _current_leave_approver_assignment():
    return LeaveApprover.query.filter_by(user_id=current_user.id).first()


def _current_leave_workflow_role():
    if (current_user.role or "").strip().lower() == "admin":
        return "admin"

    assignment = _current_leave_approver_assignment()
    if assignment and assignment.approver_role in LEAVE_APPROVER_ROLE_KEYS:
        return assignment.approver_role
    return ""


def _current_user_can_manage_leaves():
    return _current_leave_workflow_role() in {"admin", *LEAVE_APPROVER_ROLE_KEYS}


def _leave_submitter_department(leave_request):
    submitter = leave_request.submitted_by_user if leave_request else None
    if not submitter or not submitter.employee:
        return ""
    return _normalize_department_name(submitter.employee.department)


def _department_leave_approver_can_access(leave_request):
    assignment = _current_leave_approver_assignment()
    if not assignment or assignment.approver_role != "department":
        return False
    manager_department = _normalize_department_name(assignment.department_name)
    request_department = _leave_submitter_department(leave_request)
    return bool(manager_department and request_department and manager_department == request_department)


def _scope_leave_query_for_current_user(leave_query):
    current_role = _current_leave_workflow_role()
    if current_role == "department":
        assignment = _current_leave_approver_assignment()
        manager_department = _normalize_department_name(assignment.department_name if assignment else "")
        if not manager_department:
            return leave_query.filter(LeaveRequest.id.is_(None))
        return (
            leave_query
            .join(LeaveRequest.submitted_by_user)
            .join(User.employee)
            .filter(func.lower(func.trim(Employee.department)) == manager_department)
        )
    return leave_query


def _format_employee_no(hired_date, sequence):
    date_part = hired_date.strftime("%m%d%Y") if hired_date else "00000000"
    return f"{date_part}-{sequence:02d}"


def _next_employee_no_for_date(hired_date, exclude_employee_id=None):
    prefix = hired_date.strftime("%m%d%Y") if hired_date else "00000000"
    query = Employee.query.with_entities(Employee.employee_no).filter(
        Employee.employee_no.like(f"{prefix}-%")
    )
    if exclude_employee_id is not None:
        query = query.filter(Employee.id != exclude_employee_id)

    max_sequence = 0
    for value, in query.all():
        if not value:
            continue
        try:
            max_sequence = max(max_sequence, int(str(value).rsplit("-", 1)[1]))
        except (IndexError, ValueError):
            continue
    return _format_employee_no(hired_date, max_sequence + 1)


def _should_refresh_employee_no(employee_no, hired_date):
    normalized_employee_no = (str(employee_no or "").strip()).lower()
    return (
        not normalized_employee_no
        or normalized_employee_no in {"none", "n/a", "null"}
        or normalized_employee_no.startswith("00000000-")
    )


def _compute_age_from_birth_date(birth_date):
    if not birth_date:
        return None

    today = date.today()
    age = today.year - birth_date.year - (
        (today.month, today.day) < (birth_date.month, birth_date.day)
    )
    return age if age >= 0 else None


def _is_at_least_one_year_from_hired_date(hired_date):
    if not hired_date:
        return False

    today = date.today()
    years = today.year - hired_date.year
    if (today.month, today.day) < (hired_date.month, hired_date.day):
        years -= 1
    return years >= 1


EMPLOYEE_TEXT_FIELDS = [
    "birth_place",
    "nationality",
    "height",
    "weight",
    "civil_status",
    "house_phone",
    "social_media_type",
    "social_media_detail",
    "permanent_address",
    "elementary_school",
    "elementary_year_attended",
    "secondary_school",
    "secondary_year_attended",
    "college_school",
    "college_year_attended",
    "college_course",
    "year_graduated",
    "father_name",
    "father_occupation",
    "mother_maiden_name",
    "mother_occupation",
    "no_of_siblings",
    "sibling_birth_order",
    "spouse_full_name",
    "spouse_school",
    "spouse_course_degree",
    "spouse_occupation",
    "bank_type",
    "valid_id_type",
]

EMPLOYEE_INT_FIELDS = [
    "spouse_age",
    "no_of_children",
    "no_of_male_children",
    "no_of_female_children",
]

EMPLOYEE_DATE_FIELDS = [
    "spouse_birth_date",
]


def _apply_extended_employee_fields(employee):
    for field in EMPLOYEE_TEXT_FIELDS:
        setattr(employee, field, (request.form.get(field) or "").strip())

    for field in EMPLOYEE_INT_FIELDS:
        setattr(employee, field, _parse_int(request.form.get(field)))

    for field in EMPLOYEE_DATE_FIELDS:
        setattr(employee, field, _parse_date(request.form.get(field)))

    employee.children_details = _parse_children_details()


def _parse_children_details():
    names = request.form.getlist("child_full_name[]")
    ages = request.form.getlist("child_age[]")
    birth_dates = request.form.getlist("child_birth_date[]")
    schools = request.form.getlist("child_school[]")
    school_levels = request.form.getlist("child_school_level[]")
    occupations = request.form.getlist("child_occupation[]")

    children = []
    total = max(
        len(names),
        len(ages),
        len(birth_dates),
        len(schools),
        len(school_levels),
        len(occupations),
    )

    for index in range(total):
        child = {
            "label": f"Child {index + 1}",
            "full_name": names[index].strip() if index < len(names) else "",
            "age": _parse_int(ages[index]) if index < len(ages) else None,
            "birth_date": birth_dates[index].strip() if index < len(birth_dates) else "",
            "school": schools[index].strip() if index < len(schools) else "",
            "school_level": school_levels[index].strip() if index < len(school_levels) else "",
            "occupation": occupations[index].strip() if index < len(occupations) else "",
        }
        if any(value for key, value in child.items() if key != "label"):
            children.append(child)

    return children


@admin.route("/employees")
@roles_required("admin", "hr")
def employees():
    search = (request.args.get("search") or "").strip()
    status = (request.args.get("status") or "").strip()
    company = (request.args.get("company") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 10

    employee_query = Employee.query

    if search:
        search_pattern = f"%{search}%"
        employee_query = employee_query.filter(
            or_(
                Employee.first_name.ilike(search_pattern),
                Employee.middle_name.ilike(search_pattern),
                Employee.last_name.ilike(search_pattern),
                Employee.email.ilike(search_pattern),
                Employee.employee_no.ilike(search_pattern),
                Employee.department.ilike(search_pattern),
                Employee.position.ilike(search_pattern),
                Employee.company.ilike(search_pattern),
            )
        )

    # Default: exclude Pending employees unless explicitly filtered
    if status:
        employee_query = employee_query.filter(
            or_(
                Employee.employment_status.ilike(status),
                Employee.status.ilike(status),
            )
        )
    else:
        employee_query = employee_query.filter(
            Employee.employment_status != "Pending",
            Employee.status != "Pending",
        )

    if company:
        employee_query = employee_query.filter(Employee.company == company)

    employee_pagination = employee_query.order_by(Employee.id.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )
    employee_items = employee_pagination.items
    company_items = Company.query.order_by(Company.company_name.asc()).all()
    department_items = Department.query.order_by(Department.department_name.asc()).all()
    add_employee_form = session.pop("add_employee_form", {})

    return render_template(
        "admin/employees.html",
        employees=employee_items,
        companies=company_items,
        departments=department_items,
        add_employee_form=add_employee_form,
        employee_pagination=employee_pagination,
        employee_filters={
            "search": search,
            "status": status,
            "company": company,
        },
    )


@admin.route("/companies")
@roles_required("admin", "hr")
def companies():
    search = (request.args.get("search") or "").strip()
    contact_filter = (request.args.get("contact_filter") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 10

    company_query = Company.query

    if search:
        search_pattern = f"%{search}%"
        company_query = company_query.filter(
            or_(
                Company.company_name.ilike(search_pattern),
                Company.contact_number.ilike(search_pattern),
                Company.address.ilike(search_pattern),
            )
        )

    if contact_filter == "with_contact":
        company_query = company_query.filter(
            Company.contact_number.isnot(None),
            Company.contact_number != "",
        )
    elif contact_filter == "without_contact":
        company_query = company_query.filter(
            or_(
                Company.contact_number.is_(None),
                Company.contact_number == "",
            )
        )

    company_pagination = company_query.order_by(Company.id.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )
    company_items = company_pagination.items

    return render_template(
        "admin/companies.html",
        companies=company_items,
        company_pagination=company_pagination,
        company_filters={
            "search": search,
            "contact_filter": contact_filter,
        },
    )

# from sqlalchemy import or_

@admin.route("/dashboard")
@roles_required("admin", "hr")
def dashboard():
    # Totals
    total_employees = Employee.query.count()
    total_companies = Company.query.count()
    total_departments = Department.query.count()

    # Pending employees
    pending_employees = Employee.query.filter(
        or_(
            Employee.status == "Pending",
            Employee.employment_status == "Pending"
        )
    ).count()

    # Status breakdown
    status_breakdown = (
        db.session.query(Employee.employment_status, func.count(Employee.id))
        .group_by(Employee.employment_status)
        .all()
    )
    status_counts = {s: c for s, c in status_breakdown if s}

    # Recent employees (last 6)
    recent_employees = (
        Employee.query.order_by(Employee.id.desc()).limit(6).all()
    )

    # Department distribution
    dept_breakdown = (
        db.session.query(Employee.department, func.count(Employee.id))
        .group_by(Employee.department)
        .all()
    )
    dept_counts = [
        {"name": d, "count": c}
        for d, c in dept_breakdown
        if d and d.strip()
    ]
    dept_counts.sort(key=lambda x: x["count"], reverse=True)

    # Company distribution
    company_breakdown = (
        db.session.query(Employee.company, func.count(Employee.id))
        .group_by(Employee.company)
        .all()
    )
    company_counts = [
        {"name": c, "count": n}
        for c, n in company_breakdown
        if c and c.strip()
    ]
    company_counts.sort(key=lambda x: x["count"], reverse=True)

    return render_template(
        "admin/dashboard.html",
        total_employees=total_employees,
        total_companies=total_companies,
        total_departments=total_departments,
        pending_employees=pending_employees,
        status_counts=status_counts,
        recent_employees=recent_employees,
        dept_counts=dept_counts,
        company_counts=company_counts,
    )


@admin.route('/register_employee')
def register_employee():
    companies = Company.query.order_by(Company.company_name.asc()).all()
    departments = Department.query.order_by(Department.department_name.asc()).all()
    return render_template('admin/register_employee.html', companies=companies, departments=departments)

@admin.route("/add-employee", methods=["POST"])
def add_employee():

    # Always default status
    employment_status = "Pending"

    # Parse birthdate and compute age
    birth_date = _parse_date(request.form.get("birth_date"))
    age = _compute_age_from_birth_date(birth_date)

    # Get name fields (ADDED FOR DUPLICATE CHECK ONLY)
    first_name = (request.form.get("first_name") or "").strip()
    last_name = (request.form.get("last_name") or "").strip()

    # Duplicate check added for matching first and last names.
    existing_employee = Employee.query.filter(
        Employee.first_name.ilike(first_name),
        Employee.last_name.ilike(last_name)
    ).first()

    if existing_employee:
        flash("Employee with the same First Name and Last Name already exists!", "error")
        return redirect(url_for("admin.employees"))

    # Photo upload
    photo_file = request.files.get("photo")
    employee_photo_path = (
        _save_employee_photo(photo_file)[0]
        if photo_file and photo_file.filename
        else None
    )
    hired_date = _parse_date(request.form.get("hired_date"))
    employee_no = _next_employee_no_for_date(hired_date)

    # Create employee (ONLY fields from your HTML form)
    new_employee = Employee(
        # Personal Info
        first_name=request.form.get("first_name"),
        middle_name=request.form.get("middle_name"),
        last_name=request.form.get("last_name"),
        suffix=request.form.get("suffix"),
        age=age,
        gender=request.form.get("gender"),
        religion=request.form.get("religion"),
        birth_date=birth_date,
        educational_attainment=request.form.get("educational_attainment"),

        # Contact Info
        email=request.form.get("email"),
        phone=request.form.get("phone"),
        zipCode=request.form.get("zipCode"),
        present_address=request.form.get("present_address"),

        # Employment Info
        employee_no=employee_no,
        company=(request.form.get("company") or "").strip(),
        department=(request.form.get("department") or "").strip(),
        hired_date=hired_date,
        employee_type=request.form.get("employee_type"),

        # Government IDs & Bank
        sss_no=(request.form.get("sss_no") or "").strip(),
        philhealth_no=(request.form.get("philhealth_no") or "").strip(),
        pagibig_no=(request.form.get("pagibig_no") or "").strip(),
        tin_no=(request.form.get("tin_no") or "").strip(),
        valid_id_no=(request.form.get("valid_id_no") or "").strip(),
        account_no=(request.form.get("account_no") or "").strip(),

        # Photo
        photopath=employee_photo_path,

        # DEFAULT STATUS (NO USER INPUT)
        status=(request.form.get("employment_status") or "Active"),
        employment_status=(request.form.get("employment_status") or "Active"),
    )
    _apply_extended_employee_fields(new_employee)

    try:
        db.session.add(new_employee)
        sync_department_name(new_employee.department)
        db.session.flush()  # get new_employee.id before creating user

        # Create user account if username is provided
        reg_username = (request.form.get("reg_username") or "").strip()
        reg_password = request.form.get("reg_password") or ""
        reg_confirm = request.form.get("reg_confirm_password") or ""
        reg_role = "user"

        if reg_username:
            if len(reg_username) < 4:
                db.session.rollback()
                flash("Username must be at least 4 characters.", "error")
                return redirect(url_for("admin.employees"))
            if not reg_password or len(reg_password) < 7:
                db.session.rollback()
                flash("Password must be at least 7 characters when creating an account.", "error")
                return redirect(url_for("admin.employees"))
            if reg_password != reg_confirm:
                db.session.rollback()
                flash("Passwords do not match.", "error")
                return redirect(url_for("admin.employees"))

            existing_user = User.query.filter_by(username=reg_username).first()
            if existing_user:
                db.session.rollback()
                flash(f"Username '{reg_username}' is already taken.", "error")
                return redirect(url_for("admin.employees"))

            new_user = User(
                username=reg_username,
                password=generate_password_hash(reg_password, method="pbkdf2:sha256"),
                role=reg_role,
                employee_id=new_employee.id,
                leave_credits=7 if _is_at_least_one_year_from_hired_date(new_employee.hired_date) else 0,
                offset_credits=0.0,
            )
            db.session.add(new_user)

        db.session.commit()
        flash("Employee added successfully.", "success")

    except Exception:
        db.session.rollback()
        flash("Error saving employee.", "error")
        return redirect(url_for("admin.employees"))

    return redirect(url_for("employee.view_employee", employee_id=new_employee.id))
# EDIT EMPLOYEE
@admin.route("/edit-employee/<int:employee_id>", methods=["POST"])
@roles_required("admin", "hr")
def edit_employee(employee_id):
    employee = Employee.query.filter_by(id=employee_id).first()
    if not employee:
        msg = "Employee not found."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"success": False, "message": msg}), 404
        flash(msg, category="error")
        return redirect(url_for("admin.employees"))

    fields_to_check = {
        "email": "Email Address",
        "employee_no": "Employee ID No.",
        "biometric_no": "Biometric No.",
        "sss_no": "SSS No.",
        "philhealth_no": "PhilHealth No.",
        "pagibig_no": "Pag-IBIG No.",
        "tin_no": "TIN No.",
        "valid_id_no": "Valid ID No.",
        "account_no": "Bank Account No.",
    }

    duplicates = []
    existing_employees = Employee.query.filter(Employee.id != employee.id).all()

    for field, label in fields_to_check.items():
        form_val = (request.form.get(field) or "").strip()
        if not form_val:
            continue

        # Normalize form value
        norm_form = form_val.lower() if field == "email" else re.sub(r"[^a-zA-Z0-9]", "", form_val).lower()
        if not norm_form:
            continue

        # Skip duplicate check if the user didn't change this field
        raw_current = getattr(employee, field, None)
        current_val = "" if raw_current is None else str(raw_current).strip()
        norm_current = current_val.lower() if field == "email" else re.sub(r"[^a-zA-Z0-9]", "", current_val).lower()
        if norm_form == norm_current:
            continue

        for emp in existing_employees:
            emp_val = getattr(emp, field, None)
            emp_val = "" if emp_val is None else str(emp_val).strip()

            if not emp_val:
                continue

            if field == "email":
                clean_emp = emp_val.lower()
            else:
                clean_emp = re.sub(r"[^a-zA-Z0-9]", "", emp_val).lower()

            if norm_form == clean_emp:
                duplicates.append(label)
                break

    email = (request.form.get("email") or "").strip()
    employee_no = (request.form.get("employee_no") or "").strip()
    biometric_no = (request.form.get("biometric_no") or "").strip()

    if duplicates:
        msg = f"Duplicate data found for: {', '.join(duplicates)}."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"success": False, "message": msg}), 409
        flash(msg, category="error")
        return redirect(url_for("admin.employees"))

    employment_status = (request.form.get("employment_status") or "").strip() or employee.employment_status
    company_name = (request.form.get("company") or "").strip()
    birth_date = _parse_date(request.form.get("birth_date")) or employee.birth_date
    hired_date = _parse_date(request.form.get("hired_date")) or employee.hired_date
    if _should_refresh_employee_no(employee_no, hired_date):
        employee_no = _next_employee_no_for_date(hired_date, exclude_employee_id=employee.id)

    first_name = (request.form.get("first_name") or "").strip()
    last_name = (request.form.get("last_name") or "").strip()
    department = (request.form.get("department") or "").strip()
    position = (request.form.get("position") or "").strip()

    # Process photo update if a new photo is uploaded
    photo_file = request.files.get("photo")
    if photo_file and photo_file.filename:
        employee_photo_path, _ = _save_employee_photo(photo_file)
        if employee_photo_path:
            employee.photopath = employee_photo_path  # Only update if a valid new photo exists

    # 1. Personal Information
    employee.first_name = first_name
    employee.middle_name = (request.form.get("middle_name") or "").strip()
    employee.last_name = last_name
    employee.suffix = (request.form.get("suffix") or "").strip()
    employee.birth_date = birth_date
    employee.age = _parse_int(request.form.get("age"))
    employee.gender = request.form.get("gender")
    employee.religion = (request.form.get("religion") or "").strip()
    employee.biometric_no = biometric_no
    employee.educational_attainment = request.form.get("educational_attainment")

    # 2. Contact & Address
    employee.email = email
    employee.phone = (request.form.get("phone") or "").strip()
    employee.zipCode = (request.form.get("zipCode") or "").strip()
    employee.present_address = (request.form.get("present_address") or "").strip()

    # 3. Employment Details
    employee.employee_no = employee_no
    employee.company = company_name
    employee.department = department
    employee.position = position
    employee.hired_date = hired_date
    employee.employee_type = request.form.get("employee_type")
    employee.employment_status = employment_status
    employee.status = employment_status
    sync_department_name(employee.department)

    # 4. Government IDs & Bank
    employee.sss_no = (request.form.get("sss_no") or "").strip()
    employee.philhealth_no = (request.form.get("philhealth_no") or "").strip()
    employee.pagibig_no = (request.form.get("pagibig_no") or "").strip()
    employee.tin_no = (request.form.get("tin_no") or "").strip()
    employee.valid_id_no = (request.form.get("valid_id_no") or "").strip()
    employee.account_no = (request.form.get("account_no") or "").strip()
    _apply_extended_employee_fields(employee)

    try:
        db.session.commit()
        msg = "Employee updated successfully."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"success": True, "message": msg})
        flash(msg, category="success")
    except IntegrityError:
        db.session.rollback()
        msg = "Duplicate data detected. Please use unique employee details."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"success": False, "message": msg}), 409
        flash(msg, category="error")

    return redirect(url_for("admin.employees"))


@admin.route("/add-company", methods=["POST"])
@roles_required("admin", "hr")
def add_company():
    company_name = (request.form.get("company_name") or "").strip()
    contact_number = request.form.get("contact_number")
    address = request.form.get("address")
    logo_path = _save_company_logo(request.files.get("company_logo"))

    if not company_name:
        flash("Company name is required.", category="error")
        return redirect(url_for("admin.companies"))

    new_company = Company(
        company_name=company_name,
        contact_number=contact_number,
        address=address,
        logo_path=logo_path,
    )

    try:
        db.session.add(new_company)
        db.session.commit()
        flash("Company added successfully.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash("Company could not be saved. It may already exist.", category="error")

    return redirect(url_for("admin.companies"))


@admin.route("/edit-company/<int:company_id>", methods=["POST"])
@roles_required("admin", "hr")
def edit_company(company_id):
    company = Company.query.filter_by(id=company_id).first()
    if not company:
        flash("Company not found.", category="error")
        return redirect(url_for("admin.companies"))

    company_name = (request.form.get("company_name") or "").strip()
    contact_number = (request.form.get("contact_number") or "").strip()
    address = (request.form.get("address") or "").strip()

    if not company_name:
        flash("Company name is required.", category="error")
        return redirect(url_for("admin.companies"))

    duplicate_company = Company.query.filter(
        Company.id != company.id,
        Company.company_name == company_name,
    ).first()
    if duplicate_company:
        flash("Company name already exists. Please use a different name.", category="error")
        return redirect(url_for("admin.companies"))

    company.company_name = company_name
    company.contact_number = contact_number or None
    company.address = address or None

    logo_file = request.files.get("company_logo")
    if logo_file and logo_file.filename:
        logo_path = _save_company_logo(logo_file)
        if not logo_path:
            flash("Company logo must be a JPG or PNG file.", category="error")
            return redirect(url_for("admin.companies"))
        company.logo_path = logo_path

    try:
        db.session.commit()
        flash("Company updated successfully.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash("Company could not be updated. It may already exist.", category="error")

    return redirect(url_for("admin.companies"))


@admin.route("/departments")
@roles_required("admin", "hr")
def departments():
    search = (request.args.get("search") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 10

    department_query = Department.query
    if search:
        department_query = department_query.filter(
            Department.department_name.ilike(f"%{search}%")
        )

    department_pagination = department_query.order_by(Department.department_name.asc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )
    department_items = department_pagination.items

    employee_count_rows = (
        db.session.query(Employee.department, func.count(Employee.id))
        .filter(Employee.department.isnot(None), Employee.department != "")
        .group_by(Employee.department)
        .all()
    )
    employee_counts = {
        department_name: count
        for department_name, count in employee_count_rows
    }

    return render_template(
        "admin/departments.html",
        departments=department_items,
        department_pagination=department_pagination,
        department_filters={"search": search},
        employee_counts=employee_counts,
    )


@admin.route("/departments/add", methods=["POST"])
@roles_required("admin", "hr")
def add_department():
    department_name = (request.form.get("department_name") or "").strip()
    if not department_name:
        flash("Department name is required.", category="error")
        return redirect(url_for("admin.departments"))

    existing_department = Department.query.filter(
        func.lower(Department.department_name) == department_name.lower()
    ).first()
    if existing_department:
        flash("Department already exists.", category="error")
        return redirect(url_for("admin.departments"))

    db.session.add(Department(department_name=department_name))
    try:
        db.session.commit()
        flash("Department added successfully.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash("Department already exists.", category="error")

    return redirect(url_for("admin.departments"))


@admin.route("/departments/<int:department_id>/edit", methods=["POST"])
@roles_required("admin", "hr")
def edit_department(department_id):
    department = Department.query.get_or_404(department_id)
    department_name = (request.form.get("department_name") or "").strip()
    if not department_name:
        flash("Department name is required.", category="error")
        return redirect(url_for("admin.departments"))

    duplicate_department = Department.query.filter(
        Department.id != department.id,
        func.lower(Department.department_name) == department_name.lower(),
    ).first()
    if duplicate_department:
        flash("Department already exists.", category="error")
        return redirect(url_for("admin.departments"))

    old_department_name = department.department_name
    department.department_name = department_name
    Employee.query.filter(Employee.department == old_department_name).update(
        {Employee.department: department_name},
        synchronize_session=False,
    )
    EsarfApprover.query.filter(EsarfApprover.department_name == old_department_name).update(
        {EsarfApprover.department_name: department_name},
        synchronize_session=False,
    )

    try:
        db.session.commit()
        flash("Department updated successfully.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash("Department could not be updated.", category="error")

    return redirect(url_for("admin.departments"))


@admin.route('/users', methods=['GET'])
@roles_required("admin")
def users():
    user_items = User.query.order_by(User.id.desc()).all()
    return render_template('admin/users.html', users=user_items)


@admin.route("/update-user/<int:user_id>", methods=["POST"])
@roles_required("admin")
def update_user(user_id):
    user = User.query.filter_by(id=user_id).first()
    if not user:
        flash("User not found.", category="error")
        return redirect(url_for("admin.users"))

    raw_role = request.form.get("role") or ""
    role_key = " ".join(
        raw_role.strip().lower().replace("_", " ").replace("-", " ").replace(",", " ").split()
    )
    role_aliases = {
        "admin": "admin",
        "user": "user",
        "hr": "hr",
        "timekeeper": "timekeeper",
        "general manager": "general manager",
        "gm": "general manager",
        "operation": "operation",
        "operations": "operation",
        "dept manager": "dept manager",
        "department manager": "dept manager",
    }
    role = role_aliases.get(role_key)
    if not role:
        flash("Invalid user role.", category="error")
        return redirect(url_for("admin.users"))

    user.role = role

    if user.employee:
        employment_status = (request.form.get("employment_status") or "").strip()
        if employment_status:
            user.employee.employment_status = employment_status
            user.employee.status = employment_status

    try:
        db.session.commit()
        flash("User updated successfully.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash("Could not update user. Please try again.", category="error")

    return redirect(url_for("admin.users"))


@admin.route('/esarf_requests', methods=['GET'])
@login_required
def esarf_requests():
    if not (_current_user_can_manage_esarf() or _current_user_can_manage_leaves()):
        flash("You do not have permission to access that page.", category="error")
        return redirect(url_for("views.home"))

    esarf_request_query = EsarfRequest.query
    current_role = _current_esarf_workflow_role()
    if _current_user_can_manage_esarf():
        esarf_request_query = _scope_esarf_query_for_current_user(esarf_request_query)
        if current_role == "dept manager":
            esarf_request_query = esarf_request_query.filter(
                EsarfRequest.status == ESARF_STATUS_PENDING
            )
        elif current_role == "operation":
            esarf_request_query = esarf_request_query.filter(
                EsarfRequest.status.in_(
                    [
                        ESARF_STATUS_DEPT_MGR_APPROVED,
                    ]
                )
            )
        elif current_role == "general manager":
            esarf_request_query = esarf_request_query.filter(
                EsarfRequest.status.in_(
                    [
                        ESARF_STATUS_DEPT_MGR_OPS_APPROVED,
                    ]
                )
            )
        esarf_request_items = esarf_request_query.order_by(EsarfRequest.id.desc()).all()
        if current_role == "general manager":
            esarf_request_items = [
                req for req in esarf_request_items if _requires_general_manager_approval(req.submitted_by_user)
            ]
    else:
        esarf_request_items = []

    leave_current_role = _current_leave_workflow_role()
    if _current_user_can_manage_leaves():
        leave_request_query = _scope_leave_query_for_current_user(LeaveRequest.query)
        if leave_current_role in {"department", "hr"}:
            leave_request_query = leave_request_query.filter(
                LeaveRequest.status == LEAVE_STATUS_PENDING
            )
        elif leave_current_role == "operation":
            leave_request_query = leave_request_query.filter(
                LeaveRequest.status == LEAVE_STATUS_DEPT_HR_APPROVED
            )
        elif leave_current_role == "general manager":
            leave_request_query = leave_request_query.filter(
                LeaveRequest.status == LEAVE_STATUS_DEPT_HR_OPS_APPROVED
            )
        leave_request_items = leave_request_query.order_by(LeaveRequest.id.desc()).all()
        if leave_current_role == "general manager":
            leave_request_items = [
                req for req in leave_request_items if _requires_general_manager_approval(req.submitted_by_user)
            ]
    else:
        leave_request_items = []

    esarf_counts = {"total": 0, "pending": 0, "approved": 0, "rejected": 0, "returned": 0}
    for req in esarf_request_items:
        esarf_counts["total"] += 1
        status = (req.status or ESARF_STATUS_PENDING).strip()
        if status == ESARF_STATUS_PENDING:
            esarf_counts["pending"] += 1
        elif status == ESARF_STATUS_APPROVED:
            esarf_counts["approved"] += 1
        elif status == ESARF_STATUS_REJECTED:
            esarf_counts["rejected"] += 1
        else:
            esarf_counts["returned"] += 1

    leave_counts = {"total": 0, "pending": 0, "approved": 0, "rejected": 0, "returned": 0}
    for leave in leave_request_items:
        leave_counts["total"] += 1
        status = (leave.status or LEAVE_STATUS_PENDING).strip()
        if status == LEAVE_STATUS_PENDING:
            leave_counts["pending"] += 1
        elif status == LEAVE_STATUS_APPROVED:
            leave_counts["approved"] += 1
        elif status == LEAVE_STATUS_REJECTED:
            leave_counts["rejected"] += 1
        else:
            leave_counts["returned"] += 1

    return render_template(
        'admin/esarf_requests.html',
        esarf_requests=esarf_request_items,
        leave_requests=leave_request_items,
        esarf_current_role=current_role,
        leave_current_role=leave_current_role,
        esarf_counts=esarf_counts,
        leave_counts=leave_counts,
        initial_request_type=(request.args.get("type") or "all").strip().lower(),
    )


@admin.route("/esarf_requests/<int:esarf_id>/status", methods=["POST"])
@login_required
def update_esarf_status(esarf_id):
    if not _current_user_can_manage_esarf():
        flash("You do not have permission to update this request.", category="error")
        return redirect(url_for("views.home"))

    esarf_request = EsarfRequest.query.filter_by(id=esarf_id).first()
    if not esarf_request:
        flash("ESARF request not found.", category="error")
        return redirect(url_for("admin.esarf_requests"))

    current_role = _current_esarf_workflow_role()
    action = (request.form.get("action") or "").strip().lower()
    previous_status = esarf_request.status

    if current_role == "dept manager" and not _department_manager_can_access_esarf(esarf_request):
        flash("You can only view or approve ESARF requests from your own department.", category="error")
        return redirect(url_for("admin.esarf_requests"))

    # Backward-compatible fallback for old forms that still submit "status".
    legacy_status = (request.form.get("status") or "").strip().title()
    if not action and legacy_status == ESARF_STATUS_APPROVED:
        if current_role == "dept manager":
            action = "dept_manager_approve"
        elif current_role == "operation":
            action = "operation_approve"
        elif current_role == "general manager":
            action = "general_manager_approve"
        else:
            action = "approve"
    elif not action and legacy_status == ESARF_STATUS_REJECTED:
        action = "reject"

    role_labels = {
        "admin": "Admin",
        "timekeeper": "Timekeeper",
        "dept manager": "Dept Manager",
        "operation": "Operations",
        "general manager": "General Manager",
    }
    approver_label = role_labels.get(current_role, current_role.title())

    if current_role == "dept manager":
        if action == "reject":
            if esarf_request.status != ESARF_STATUS_PENDING:
                flash("Only pending requests can be declined.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            reject_reason = (request.form.get("reject_reason") or "").strip()
            if not reject_reason:
                flash("Decline reason is required.", category="error")
                return redirect(url_for("admin.esarf_requests"))

            esarf_request.status = ESARF_STATUS_REJECTED
            esarf_request.declined_reason = f"Dept Manager: {reject_reason}"
            success_message = f"ESARF request #{esarf_request.id} declined by Dept Manager. Reason: {reject_reason}"
        elif action == "dept_manager_approve":
            if esarf_request.status in {ESARF_STATUS_REJECTED, ESARF_STATUS_APPROVED}:
                flash("This request can no longer be department-manager approved.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            if esarf_request.status in {ESARF_STATUS_DEPT_MGR_APPROVED, ESARF_STATUS_DEPT_MGR_OPS_APPROVED}:
                flash("Department manager approval is already recorded.", category="info")
                return redirect(url_for("admin.esarf_requests"))

            esarf_request.status = ESARF_STATUS_DEPT_MGR_APPROVED
            success_message = (
                f"ESARF request #{esarf_request.id}: Department Manager approval recorded. "
                "Waiting for remaining signatories."
            )
        else:
            flash(
                "Department manager can approve as first signatory or decline pending requests.",
                category="error",
            )
            return redirect(url_for("admin.esarf_requests"))
    elif current_role == "operation":
        if action == "reject":
            if esarf_request.status != ESARF_STATUS_DEPT_MGR_APPROVED:
                flash("Only requests awaiting operations approval can be declined.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            reject_reason = (request.form.get("reject_reason") or "").strip()
            if not reject_reason:
                flash("Decline reason is required.", category="error")
                return redirect(url_for("admin.esarf_requests"))

            esarf_request.status = ESARF_STATUS_REJECTED
            esarf_request.declined_reason = f"Operations: {reject_reason}"
            success_message = f"ESARF request #{esarf_request.id} declined by Operations. Reason: {reject_reason}"
        elif action == "operation_approve":
            if esarf_request.status in {ESARF_STATUS_REJECTED, ESARF_STATUS_APPROVED}:
                flash("This request can no longer be operation-approved.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            if esarf_request.status == ESARF_STATUS_PENDING:
                flash("Department manager approval is required before operation approval.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            if esarf_request.status == ESARF_STATUS_DEPT_MGR_OPS_APPROVED:
                flash("Operation approval is already recorded.", category="info")
                return redirect(url_for("admin.esarf_requests"))
            if esarf_request.status != ESARF_STATUS_DEPT_MGR_APPROVED:
                flash("Current request status is not valid for operation approval.", category="error")
                return redirect(url_for("admin.esarf_requests"))

            if _requires_general_manager_approval(esarf_request.submitted_by_user):
                esarf_request.status = ESARF_STATUS_DEPT_MGR_OPS_APPROVED
                success_message = (
                    f"ESARF request #{esarf_request.id}: Operations approval recorded. "
                    "Waiting for remaining signatory."
                )
            else:
                esarf_request.status = ESARF_STATUS_APPROVED
                success_message = (
                    f"ESARF request #{esarf_request.id}: Operations approval recorded. "
                    "Request is now Approved."
                )
        else:
            flash("Operation role can approve or decline requests awaiting second-signatory action.", category="error")
            return redirect(url_for("admin.esarf_requests"))
    elif current_role == "general manager":
        if not _requires_general_manager_approval(esarf_request.submitted_by_user):
            flash("General Manager approval is not required for this request.", category="info")
            return redirect(url_for("admin.esarf_requests"))
        if action == "reject":
            if esarf_request.status != ESARF_STATUS_DEPT_MGR_OPS_APPROVED:
                flash("Only requests awaiting general manager approval can be declined.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            reject_reason = (request.form.get("reject_reason") or "").strip()
            if not reject_reason:
                flash("Decline reason is required.", category="error")
                return redirect(url_for("admin.esarf_requests"))

            esarf_request.status = ESARF_STATUS_REJECTED
            esarf_request.declined_reason = f"General Manager: {reject_reason}"
            success_message = f"ESARF request #{esarf_request.id} declined by General Manager. Reason: {reject_reason}"
        elif action == "general_manager_approve":
            if esarf_request.status in {ESARF_STATUS_REJECTED, ESARF_STATUS_APPROVED}:
                flash("This request can no longer be general-manager approved.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            if esarf_request.status != ESARF_STATUS_DEPT_MGR_OPS_APPROVED:
                flash("Operations approval is required before general manager approval.", category="error")
                return redirect(url_for("admin.esarf_requests"))

            esarf_request.status = ESARF_STATUS_APPROVED
            success_message = f"ESARF request #{esarf_request.id}: General Manager approval recorded. Request is now Approved."
        else:
            flash("General manager can approve or decline requests awaiting final-signatory action.", category="error")
            return redirect(url_for("admin.esarf_requests"))
    elif current_role in {"admin", "timekeeper"}:
        if action == "approve":
            if esarf_request.status != ESARF_STATUS_PENDING:
                flash("Only pending requests can be approved.", category="error")
                return redirect(url_for("admin.esarf_requests"))

            esarf_request.status = ESARF_STATUS_APPROVED
            success_message = f"ESARF request #{esarf_request.id} approved by {approver_label}."
        elif action == "reject":
            if esarf_request.status != ESARF_STATUS_PENDING:
                flash("Only pending requests can be declined.", category="error")
                return redirect(url_for("admin.esarf_requests"))
            reject_reason = (request.form.get("reject_reason") or "").strip()
            if not reject_reason:
                flash("Decline reason is required.", category="error")
                return redirect(url_for("admin.esarf_requests"))

            esarf_request.status = ESARF_STATUS_REJECTED
            esarf_request.declined_reason = f"{approver_label}: {reject_reason}"
            success_message = f"ESARF request #{esarf_request.id} declined by {approver_label}. Reason: {reject_reason}"
        else:
            flash("Admin approvers can only approve or reject pending ESARF requests.", category="error")
            return redirect(url_for("admin.esarf_requests"))
    else:
        flash("You do not have permission to update this request.", category="error")
        return redirect(url_for("admin.esarf_requests"))

    try:
        moved_to_approved = previous_status != ESARF_STATUS_APPROVED and esarf_request.status == ESARF_STATUS_APPROVED
        if moved_to_approved:
            submitter = esarf_request.submitted_by_user
            submitter.offset_credits = float(submitter.offset_credits or 0)

            transaction_csv = esarf_request.transaction_types or ""
            tokens = [token.strip() for token in transaction_csv.split(",") if token.strip()]
            if "Use Offset" in tokens:
                if submitter.offset_credits < float(esarf_request.total_hours or 0):
                    flash(
                        f"Cannot approve: insufficient offset credits. Available {submitter.offset_credits:.2f} hrs, "
                        f"requested {float(esarf_request.total_hours or 0):.2f} hrs.",
                        category="error",
                    )
                    db.session.rollback()
                    return redirect(url_for("admin.esarf_requests"))
                submitter.offset_credits = max(0.0, submitter.offset_credits - float(esarf_request.total_hours or 0))
            elif "Offset" in tokens:
                submitter.offset_credits = submitter.offset_credits + float(esarf_request.total_hours or 0)

        notification_category = "approved" if esarf_request.status == ESARF_STATUS_APPROVED else (
            "rejected" if esarf_request.status == ESARF_STATUS_REJECTED else "info"
        )
        if esarf_request.status == ESARF_STATUS_REJECTED:
            notification_title = f"Declined by {approver_label}"
            notification_message = "Your ESARF was declined."
        elif esarf_request.status == ESARF_STATUS_APPROVED:
            notification_title = f"Approved by {approver_label}"
            notification_message = "Your ESARF was approved."
        elif action == "operation_approve":
            notification_title = "Approved by Operations"
            if esarf_request.status == ESARF_STATUS_APPROVED:
                notification_message = "Your ESARF was approved."
            else:
                notification_message = "Moved to General Manager."
        else:
            notification_title = "Approved by Dept Manager"
            notification_message = "Moved to the next approver."

        create_notification(
            esarf_request.submitted_by_user_id,
            notification_title,
            notification_message,
            category=notification_category,
            link_url=url_for("employee.esarf_requests"),
        )
        db.session.commit()
        flash(success_message, category="success")
    except Exception:
        db.session.rollback()
        flash("Unable to update ESARF request status.", category="error")

    return redirect(url_for("admin.esarf_requests"))


@admin.route('/leave_requests', methods=['GET'])
@login_required
def leave_requests():
    return redirect(url_for("admin.esarf_requests"))


@admin.route('/leave_requests/<int:leave_id>/status', methods=['POST'])
@login_required
def update_leave_status(leave_id):
    if not _current_user_can_manage_leaves():
        flash('You do not have permission to update leave requests.', category='error')
        return redirect(url_for('views.home'))

    leave_request = LeaveRequest.query.get_or_404(leave_id)
    current_role = _current_leave_workflow_role()
    if current_role == "department" and not _department_leave_approver_can_access(leave_request):
        flash('You can only approve leave requests from your own department.', category='error')
        return redirect(url_for('admin.esarf_requests'))

    status = (request.form.get('status') or '').strip().title()
    if status not in {LEAVE_STATUS_APPROVED, LEAVE_STATUS_REJECTED}:
        flash('Please choose a valid leave status.', category='error')
        return redirect(url_for('admin.esarf_requests'))

    leave_status = (leave_request.status or LEAVE_STATUS_PENDING).strip()
    if leave_status not in {
        LEAVE_STATUS_PENDING,
        LEAVE_STATUS_DEPT_HR_APPROVED,
        LEAVE_STATUS_DEPT_HR_OPS_APPROVED,
    }:
        flash('This leave request can no longer be updated.', category='error')
        return redirect(url_for('admin.esarf_requests'))

    if current_role in {"department", "hr"}:
        if leave_status != LEAVE_STATUS_PENDING:
            flash('First approval can only be done while leave is pending.', category='error')
            return redirect(url_for('admin.esarf_requests'))
        if status == LEAVE_STATUS_APPROVED:
            leave_request.status = LEAVE_STATUS_DEPT_HR_APPROVED
            notification_title = "Leave pre-approved"
            notification_message = "Your leave request passed Department/HR approval and moved to Operations."
            notification_category = "info"
        else:
            leave_request.status = LEAVE_STATUS_REJECTED
            notification_title = "Leave rejected"
            notification_message = f"Your {leave_request.leave_category} leave was rejected."
            notification_category = "rejected"
    elif current_role == "operation":
        if leave_status != LEAVE_STATUS_DEPT_HR_APPROVED:
            flash('Operations can only act after Department/HR approval.', category='error')
            return redirect(url_for('admin.esarf_requests'))
        if status == LEAVE_STATUS_APPROVED:
            if _requires_general_manager_approval(leave_request.submitted_by_user):
                leave_request.status = LEAVE_STATUS_DEPT_HR_OPS_APPROVED
                notification_title = "Leave approved by Operations"
                notification_message = (
                    f"Your {leave_request.leave_category} leave passed Operations approval and moved to General Manager."
                )
                notification_category = "info"
            else:
                leave_request.status = LEAVE_STATUS_APPROVED
                deducted_days = _apply_leave_credit_deduction(leave_request)
                notification_title = "Leave approved"
                notification_message = f"Your {leave_request.leave_category} leave was approved."
                notification_category = "approved"
        else:
            leave_request.status = LEAVE_STATUS_REJECTED
            notification_title = "Leave rejected"
            notification_message = f"Your {leave_request.leave_category} leave was rejected by Operations."
            notification_category = "rejected"
    elif current_role == "general manager":
        if not _requires_general_manager_approval(leave_request.submitted_by_user):
            flash('General Manager approval is not required for this request.', category='info')
            return redirect(url_for('admin.esarf_requests'))
        if leave_status != LEAVE_STATUS_DEPT_HR_OPS_APPROVED:
            flash('General Manager can only act after Operations approval.', category='error')
            return redirect(url_for('admin.esarf_requests'))
        if status == LEAVE_STATUS_APPROVED:
            leave_request.status = LEAVE_STATUS_APPROVED
            deducted_days = _apply_leave_credit_deduction(leave_request)
            notification_title = "Leave approved"
            notification_message = f"Your {leave_request.leave_category} leave was approved."
            notification_category = "approved"
        else:
            leave_request.status = LEAVE_STATUS_REJECTED
            notification_title = "Leave rejected"
            notification_message = f"Your {leave_request.leave_category} leave was rejected by General Manager."
            notification_category = "rejected"
    elif current_role == "admin":
        leave_request.status = status
        deducted_days = 0
        if status == LEAVE_STATUS_APPROVED:
            deducted_days = _apply_leave_credit_deduction(leave_request)
        notification_title = "Leave approved" if status == LEAVE_STATUS_APPROVED else "Leave rejected"
        notification_message = (
            f"Your {leave_request.leave_category} leave was approved."
            if status == LEAVE_STATUS_APPROVED else
            f"Your {leave_request.leave_category} leave was rejected."
        )
        notification_category = "approved" if status == LEAVE_STATUS_APPROVED else "rejected"
    else:
        flash('You do not have permission to update this request.', category='error')
        return redirect(url_for('admin.esarf_requests'))
    try:
        create_notification(
            leave_request.submitted_by_user_id,
            notification_title,
            notification_message,
            category=notification_category,
            link_url=url_for("employee.leaves"),
        )
        db.session.commit()
        if leave_request.status == LEAVE_STATUS_APPROVED:
            flash(
                f'Leave request #{leave_id} updated to {leave_request.status}. '
                f'Leave credits deducted: {deducted_days} day(s).',
                category='success',
            )
        else:
            flash(f'Leave request #{leave_id} updated to {leave_request.status}.', category='success')
    except Exception:
        db.session.rollback()
        flash('Unable to update leave request status.', category='error')

    return redirect(url_for('admin.esarf_requests'))


@admin.route('/perk_requests', methods=['GET'])
@login_required
def perk_requests():
    # Allow admin, hr, or assigned perk approvers
    is_approver = PerkApprover.query.filter_by(user_id=current_user.id).first()
    if not is_approver and current_user.role not in ('admin', 'hr'):
        flash('You do not have permission to view perk requests.', category='error')
        return redirect(url_for('views.home'))

    # Build combined list
    all_requests = []
    for d in DiscountRequest.query.order_by(DiscountRequest.created_at.desc()).all():
        all_requests.append({
            'id': d.id,
            'type': 'discount',
            'status': d.status,
            'product_name': d.product_name,
            'quantity': d.quantity,
            'price': d.price,
            'amount': d.amount,
            'discounted_amount': d.discounted_amount,
            'total_amount': d.discounted_amount,
            'transaction_date': d.transaction_date,
            'created_at': d.created_at,
            'declined_reason': d.declined_reason,
            'submitted_by_user': d.submitted_by_user,
        })
    for c in ProductChargeRequest.query.order_by(ProductChargeRequest.created_at.desc()).all():
        all_requests.append({
            'id': c.id,
            'type': 'charge',
            'status': c.status,
            'product_name': c.product_name,
            'quantity': c.quantity,
            'price': c.price,
            'amount': c.total_amount,
            'discounted_amount': None,
            'total_amount': c.total_amount,
            'transaction_date': c.transaction_date,
            'created_at': c.created_at,
            'declined_reason': c.declined_reason,
            'submitted_by_user': c.submitted_by_user,
        })

    # Sort by created_at desc
    all_requests.sort(key=lambda x: x['created_at'] if x['created_at'] else date.min, reverse=True)

    # Counters
    count_all = len(all_requests)
    count_pending = sum(1 for r in all_requests if r['status'] == 'Pending')
    count_approved = sum(1 for r in all_requests if r['status'] == 'Approved')
    count_rejected = sum(1 for r in all_requests if r['status'] == 'Rejected')

    # Filters
    filter_type = request.args.get('type', '') or request.args.get('type_mobile', '')
    filter_status = request.args.get('status', '')
    filter_search = request.args.get('search', '').strip().lower()
    filter_date_from = request.args.get('date_from', '').strip()
    filter_date_to = request.args.get('date_to', '').strip()

    # Parse date range
    date_from = None
    date_to = None
    try:
        if filter_date_from:
            date_from = datetime.strptime(filter_date_from, '%Y-%m-%d').date()
    except ValueError:
        pass
    try:
        if filter_date_to:
            date_to = datetime.strptime(filter_date_to, '%Y-%m-%d').date()
    except ValueError:
        pass

    filtered = all_requests
    if filter_type:
        filtered = [r for r in filtered if r['type'] == filter_type]
    if filter_status:
        filtered = [r for r in filtered if r['status'] == filter_status]
    if filter_search:
        filtered = [r for r in filtered if (
            filter_search in (r['product_name'] or '').lower() or
            filter_search in (r['submitted_by_user'].username or '').lower() or
            (r['submitted_by_user'].employee and filter_search in (
                (r['submitted_by_user'].employee.first_name or '').lower() + ' ' +
                (r['submitted_by_user'].employee.last_name or '').lower()
            ))
        )]
    if date_from:
        filtered = [r for r in filtered if r['created_at'] and r['created_at'].date() >= date_from]
    if date_to:
        filtered = [r for r in filtered if r['created_at'] and r['created_at'].date() <= date_to]

    # Pagination
    per_page = 5
    page = request.args.get('page', 1, type=int)
    total = len(filtered)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    page_requests = filtered[start:end]

    return render_template(
        'admin/perk_requests.html',
        is_approver=is_approver,
        requests=page_requests,
        count_all=count_all,
        count_pending=count_pending,
        count_approved=count_approved,
        count_rejected=count_rejected,
        filter_type=filter_type,
        filter_status=filter_status,
        filter_search=filter_search,
        filter_date_from=filter_date_from,
        filter_date_to=filter_date_to,
        page=page,
        total_pages=total_pages,
        total=total,
        per_page=per_page,
        start=start + 1 if total > 0 else 0,
        end=min(end, total),
    )


@admin.route('/perk_requests/export', methods=['GET'])
@login_required
def export_perk_requests():
    # Permission check
    is_approver = PerkApprover.query.filter_by(user_id=current_user.id).first()
    if not is_approver and current_user.role not in ('admin', 'hr'):
        flash('You do not have permission to export perk requests.', category='error')
        return redirect(url_for('views.home'))

    # Build combined list (same logic as perk_requests)
    all_requests = []
    for d in DiscountRequest.query.order_by(DiscountRequest.created_at.desc()).all():
        all_requests.append({
            'id': d.id,
            'type': 'discount',
            'status': d.status,
            'product_name': d.product_name,
            'quantity': d.quantity,
            'price': d.price,
            'amount': d.amount,
            'discounted_amount': d.discounted_amount,
            'total_amount': d.discounted_amount,
            'transaction_date': d.transaction_date,
            'created_at': d.created_at,
            'declined_reason': d.declined_reason,
            'submitted_by_user': d.submitted_by_user,
        })
    for c in ProductChargeRequest.query.order_by(ProductChargeRequest.created_at.desc()).all():
        all_requests.append({
            'id': c.id,
            'type': 'charge',
            'status': c.status,
            'product_name': c.product_name,
            'quantity': c.quantity,
            'price': c.price,
            'amount': c.total_amount,
            'discounted_amount': None,
            'total_amount': c.total_amount,
            'transaction_date': c.transaction_date,
            'created_at': c.created_at,
            'declined_reason': c.declined_reason,
            'submitted_by_user': c.submitted_by_user,
        })

    all_requests.sort(key=lambda x: x['created_at'] if x['created_at'] else date.min, reverse=True)

    # Apply same filters
    filter_type = request.args.get('type', '') or request.args.get('type_mobile', '')
    filter_status = request.args.get('status', '')
    filter_search = request.args.get('search', '').strip().lower()
    filter_date_from = request.args.get('date_from', '').strip()
    filter_date_to = request.args.get('date_to', '').strip()

    date_from = None
    date_to = None
    try:
        if filter_date_from:
            date_from = datetime.strptime(filter_date_from, '%Y-%m-%d').date()
    except ValueError:
        pass
    try:
        if filter_date_to:
            date_to = datetime.strptime(filter_date_to, '%Y-%m-%d').date()
    except ValueError:
        pass

    filtered = all_requests
    if filter_type:
        filtered = [r for r in filtered if r['type'] == filter_type]
    if filter_status:
        filtered = [r for r in filtered if r['status'] == filter_status]
    if filter_search:
        filtered = [r for r in filtered if (
            filter_search in (r['product_name'] or '').lower() or
            filter_search in (r['submitted_by_user'].username or '').lower() or
            (r['submitted_by_user'].employee and filter_search in (
                (r['submitted_by_user'].employee.first_name or '').lower() + ' ' +
                (r['submitted_by_user'].employee.last_name or '').lower()
            ))
        )]
    if date_from:
        filtered = [r for r in filtered if r['created_at'] and r['created_at'].date() >= date_from]
    if date_to:
        filtered = [r for r in filtered if r['created_at'] and r['created_at'].date() <= date_to]

    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Type', 'Employee', 'Product', 'Qty', 'Unit Price', 'Amount', 'Discounted/Total', 'Transaction Date', 'Date Requested', 'Status', 'Decline Reason'])

    for r in filtered:
        user = r['submitted_by_user']
        emp = user.employee if user and user.employee else None
        emp_name = (
            (emp.first_name or '') +
            (' ' + emp.middle_name[:1] + '.' if emp and emp.middle_name else '') +
            ' ' + (emp.last_name or '') +
            (' ' + emp.suffix if emp and emp.suffix else '')
        ).strip() if emp else (user.username if user else 'N/A')

        final_amount = r['discounted_amount'] if r['discounted_amount'] else r['total_amount']
        writer.writerow([
            r['id'],
            r['type'].title(),
            emp_name,
            r['product_name'],
            r['quantity'],
            f"{r['price']:.2f}",
            f"{r['amount']:.2f}",
            f"{final_amount:.2f}",
            r['transaction_date'].strftime('%Y-%m-%d') if r['transaction_date'] else '',
            r['created_at'].strftime('%Y-%m-%d %H:%M') if r['created_at'] else '',
            r['status'],
            r['declined_reason'] or '',
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=perk_requests.csv'},
    )


@admin.route('/perk_requests/<string:perk_type>/<int:perk_id>/approve', methods=['POST'])
@login_required
def approve_perk(perk_type, perk_id):
    # Check if user is admin/hr or assigned approver
    is_approver = PerkApprover.query.filter_by(user_id=current_user.id).first()
    if not is_approver and current_user.role not in ('admin', 'hr'):
        flash('You do not have permission to approve perk requests.', category='error')
        return redirect(url_for('admin.perk_requests'))

    if perk_type == 'discount':
        if is_approver and not is_approver.can_approve_discount and current_user.role not in ('admin', 'hr'):
            flash('You are not authorized to approve discount requests.', category='error')
            return redirect(url_for('admin.perk_requests'))
        perk = DiscountRequest.query.get_or_404(perk_id)
    elif perk_type == 'charge':
        if is_approver and not is_approver.can_approve_charge and current_user.role not in ('admin', 'hr'):
            flash('You are not authorized to approve charge requests.', category='error')
            return redirect(url_for('admin.perk_requests'))
        perk = ProductChargeRequest.query.get_or_404(perk_id)
    else:
        flash('Invalid perk type.', category='error')
        return redirect(url_for('admin.perk_requests'))

    if perk.status != 'Pending':
        flash('Only pending requests can be approved.', category='error')
        return redirect(url_for('admin.perk_requests'))

    perk.status = 'Approved'
    try:
        create_notification(
            perk.submitted_by_user_id,
            "Perk approved",
            f"Your {perk_type} request was approved.",
            category="approved",
            link_url=url_for("employee.perks"),
        )
        db.session.commit()
        flash(f'{perk_type.title()} request #{perk_id} approved.', category='success')
    except Exception:
        db.session.rollback()
        flash('Failed to update request status.', category='error')

    return redirect(url_for('admin.perk_requests'))


@admin.route('/perk_requests/<string:perk_type>/<int:perk_id>/decline', methods=['POST'])
@login_required
def decline_perk(perk_type, perk_id):
    # Check if user is admin/hr or assigned approver
    is_approver = PerkApprover.query.filter_by(user_id=current_user.id).first()
    if not is_approver and current_user.role not in ('admin', 'hr'):
        flash('You do not have permission to decline perk requests.', category='error')
        return redirect(url_for('admin.perk_requests'))

    if perk_type == 'discount':
        if is_approver and not is_approver.can_approve_discount and current_user.role not in ('admin', 'hr'):
            flash('You are not authorized to decline discount requests.', category='error')
            return redirect(url_for('admin.perk_requests'))
        perk = DiscountRequest.query.get_or_404(perk_id)
    elif perk_type == 'charge':
        if is_approver and not is_approver.can_approve_charge and current_user.role not in ('admin', 'hr'):
            flash('You are not authorized to decline charge requests.', category='error')
            return redirect(url_for('admin.perk_requests'))
        perk = ProductChargeRequest.query.get_or_404(perk_id)
    else:
        flash('Invalid perk type.', category='error')
        return redirect(url_for('admin.perk_requests'))

    if perk.status != 'Pending':
        flash('Only pending requests can be declined.', category='error')
        return redirect(url_for('admin.perk_requests'))

    decline_reason = (request.form.get('decline_reason') or '').strip()
    if not decline_reason:
        flash('Decline reason is required.', category='error')
        return redirect(url_for('admin.perk_requests'))

    perk.status = 'Rejected'
    perk.declined_reason = decline_reason
    try:
        create_notification(
            perk.submitted_by_user_id,
            "Perk declined",
            f"Your {perk_type} request was declined.",
            category="rejected",
            link_url=url_for("employee.perks"),
        )
        db.session.commit()
        flash(f'{perk_type.title()} request #{perk_id} declined.', category='success')
    except Exception:
        db.session.rollback()
        flash('Failed to update request status.', category='error')

    return redirect(url_for('admin.perk_requests'))


# ================= SETTINGS =================

@admin.route('/settings', methods=['GET', 'POST'])
@roles_required("admin")
def settings():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'export_report':
            report_type = (request.form.get('report_type') or '').strip().lower()
            date_from_raw = (request.form.get('date_from') or '').strip()
            date_to_raw = (request.form.get('date_to') or '').strip()

            if report_type not in {'esarf', 'leave', 'discount_charge'}:
                flash('Please select a valid report type.', category='error')
                return redirect(url_for('admin.settings'))
            if not date_from_raw or not date_to_raw:
                flash('Please select both Date From and Date To.', category='error')
                return redirect(url_for('admin.settings'))

            try:
                date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
                date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid date format.', category='error')
                return redirect(url_for('admin.settings'))

            if date_to < date_from:
                flash('Date To cannot be earlier than Date From.', category='error')
                return redirect(url_for('admin.settings'))

            def _serialize(value):
                if value is None:
                    return ''
                if isinstance(value, datetime):
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                if isinstance(value, date):
                    return value.strftime('%Y-%m-%d')
                return str(value)

            def _submitter_context(user):
                employee = user.employee if user else None
                full_name = ''
                if employee:
                    parts = [employee.first_name or '', employee.middle_name or '', employee.last_name or '', employee.suffix or '']
                    full_name = ' '.join(part for part in parts if part).strip()
                return {
                    'submitted_by_user_id': user.id if user else '',
                    'submitted_by_username': user.username if user else '',
                    'submitted_by_role': user.role if user else '',
                    'submitted_by_employee_id': employee.id if employee else '',
                    'submitted_by_employee_no': employee.employee_no if employee else '',
                    'submitted_by_full_name': full_name,
                    'submitted_by_company': employee.company if employee else '',
                    'submitted_by_department': employee.department if employee else '',
                    'submitted_by_position': employee.position if employee else '',
                }

            rows = []

            if report_type == 'esarf':
                items = EsarfRequest.query.filter(
                    func.date(EsarfRequest.created_at) >= date_from,
                    func.date(EsarfRequest.created_at) <= date_to,
                ).order_by(EsarfRequest.created_at.asc(), EsarfRequest.id.asc()).all()
                base_columns = [column.name for column in EsarfRequest.__table__.columns]
                context_columns = list(_submitter_context(None).keys())
                headers = base_columns + context_columns

                for item in items:
                    row_dict = {column: _serialize(getattr(item, column)) for column in base_columns}
                    row_dict.update({k: _serialize(v) for k, v in _submitter_context(item.submitted_by_user).items()})
                    rows.append([row_dict.get(header, '') for header in headers])

                filename = f"esarf_report_{date_from}_{date_to}.xlsx"
                report_label = "ESARF Report"

            elif report_type == 'leave':
                items = LeaveRequest.query.filter(
                    LeaveRequest.start_date >= date_from,
                    LeaveRequest.start_date <= date_to,
                ).order_by(LeaveRequest.start_date.asc(), LeaveRequest.id.asc()).all()
                base_columns = [column.name for column in LeaveRequest.__table__.columns]
                context_columns = list(_submitter_context(None).keys())
                headers = base_columns + context_columns

                for item in items:
                    row_dict = {column: _serialize(getattr(item, column)) for column in base_columns}
                    row_dict.update({k: _serialize(v) for k, v in _submitter_context(item.submitted_by_user).items()})
                    rows.append([row_dict.get(header, '') for header in headers])

                filename = f"leave_report_{date_from}_{date_to}.xlsx"
                report_label = "Leave Report"

            else:
                discount_items = DiscountRequest.query.filter(
                    DiscountRequest.transaction_date >= date_from,
                    DiscountRequest.transaction_date <= date_to,
                ).order_by(DiscountRequest.transaction_date.asc(), DiscountRequest.id.asc()).all()
                charge_items = ProductChargeRequest.query.filter(
                    ProductChargeRequest.transaction_date >= date_from,
                    ProductChargeRequest.transaction_date <= date_to,
                ).order_by(ProductChargeRequest.transaction_date.asc(), ProductChargeRequest.id.asc()).all()

                discount_columns = [column.name for column in DiscountRequest.__table__.columns]
                charge_columns = [column.name for column in ProductChargeRequest.__table__.columns]
                all_request_columns = sorted(set(discount_columns + charge_columns))
                context_columns = list(_submitter_context(None).keys())
                headers = ['request_type'] + all_request_columns + context_columns

                for item in discount_items:
                    row_dict = {'request_type': 'Discount'}
                    for column in all_request_columns:
                        row_dict[column] = _serialize(getattr(item, column, ''))
                    row_dict.update({k: _serialize(v) for k, v in _submitter_context(item.submitted_by_user).items()})
                    rows.append([row_dict.get(header, '') for header in headers])

                for item in charge_items:
                    row_dict = {'request_type': 'Charge'}
                    for column in all_request_columns:
                        row_dict[column] = _serialize(getattr(item, column, ''))
                    row_dict.update({k: _serialize(v) for k, v in _submitter_context(item.submitted_by_user).items()})
                    rows.append([row_dict.get(header, '') for header in headers])

                filename = f"discount_charge_report_{date_from}_{date_to}.xlsx"
                report_label = "Discount/Charge Report"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"

            title_row = [f"{report_label} ({date_from} to {date_to})"]
            sheet.append(title_row)
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
            sheet.append(headers)
            for row in rows:
                sheet.append(row)

            title_fill = PatternFill("solid", fgColor="0F172A")
            header_fill = PatternFill("solid", fgColor="1D4ED8")
            stripe_fill = PatternFill("solid", fgColor="F8FAFC")
            white_font = Font(color="FFFFFF", bold=True)
            title_font = Font(color="FFFFFF", bold=True, size=13)
            thin_border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )

            title_cell = sheet.cell(row=1, column=1)
            title_cell.fill = title_fill
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[1].height = 26

            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=2, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            sheet.row_dimensions[2].height = 24

            max_row = sheet.max_row
            max_col = sheet.max_column

            for row_idx in range(3, max_row + 1):
                if row_idx % 2 == 1:
                    for col_idx in range(1, max_col + 1):
                        sheet.cell(row=row_idx, column=col_idx).fill = stripe_fill
                for col_idx in range(1, max_col + 1):
                    data_cell = sheet.cell(row=row_idx, column=col_idx)
                    data_cell.border = thin_border
                    data_cell.alignment = Alignment(vertical="top", wrap_text=True)

            for col_idx in range(1, max_col + 1):
                column_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, max_row + 1):
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    value_len = len(str(value)) if value is not None else 0
                    if value_len > max_len:
                        max_len = value_len
                sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 14), 42)

            sheet.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"
            sheet.freeze_panes = "A3"

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            response = Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            return response

        if action == 'assign_esarf_approver':
            user_id = request.form.get('user_id')
            approver_role = (request.form.get('approver_role') or '').strip().lower()
            department_name = (request.form.get('department_name') or '').strip()

            if approver_role not in ESARF_APPROVER_ROLE_KEYS:
                flash('Please select a valid ESARF approver role.', category='error')
                return redirect(url_for('admin.settings'))
            if not user_id:
                flash('Please select a user.', category='error')
                return redirect(url_for('admin.settings'))

            user = User.query.get(user_id)
            if not user:
                flash('User not found.', category='error')
                return redirect(url_for('admin.settings'))
            if user.role == 'admin':
                flash('Admin users cannot be reassigned from approval settings.', category='error')
                return redirect(url_for('admin.settings'))
            if approver_role == 'dept manager':
                if not department_name:
                    flash('Please choose the department this manager can approve.', category='error')
                    return redirect(url_for('admin.settings'))
                selected_department = Department.query.filter(
                    func.lower(Department.department_name) == department_name.lower()
                ).first()
                if not selected_department:
                    flash('Selected department was not found.', category='error')
                    return redirect(url_for('admin.settings'))
                department_name = selected_department.department_name
            else:
                department_name = None

            if approver_role == 'dept manager':
                existing_role_assignment = EsarfApprover.query.filter(
                    EsarfApprover.user_id != user.id,
                    EsarfApprover.approver_role == approver_role,
                    func.lower(EsarfApprover.department_name) == department_name.lower(),
                ).first()
                if existing_role_assignment:
                    flash('That department already has a Department Manager approver.', category='error')
                    return redirect(url_for('admin.settings'))

            existing_assignment = EsarfApprover.query.filter_by(user_id=user.id).first()
            if existing_assignment:
                existing_assignment.approver_role = approver_role
                existing_assignment.department_name = department_name
            else:
                db.session.add(
                    EsarfApprover(
                        user_id=user.id,
                        approver_role=approver_role,
                        department_name=department_name,
                    )
                )

            role_label = dict(ESARF_APPROVER_ROLES)[approver_role]
            try:
                db.session.commit()
                flash(f'{user.username} assigned as {role_label}.', category='success')
            except Exception:
                db.session.rollback()
                flash('Failed to assign ESARF approver.', category='error')
            return redirect(url_for('admin.settings'))

        elif action == 'remove_esarf_approver':
            user_id = request.form.get('user_id')
            user = User.query.get(user_id)
            if not user:
                flash('User not found.', category='error')
                return redirect(url_for('admin.settings'))
            assignment = EsarfApprover.query.filter_by(user_id=user.id).first()
            if not assignment:
                flash('Selected user is not an ESARF approver.', category='error')
                return redirect(url_for('admin.settings'))

            previous_role = dict(ESARF_APPROVER_ROLES).get(
                assignment.approver_role,
                'ESARF',
            )
            db.session.delete(assignment)
            try:
                db.session.commit()
                flash(f'{user.username} removed from {previous_role} approvers.', category='success')
            except Exception:
                db.session.rollback()
                flash('Failed to remove ESARF approver.', category='error')
            return redirect(url_for('admin.settings'))

        elif action == 'assign_leave_approver':
            user_id = request.form.get('user_id')
            approver_role = (request.form.get('approver_role') or '').strip().lower()
            department_name = (request.form.get('department_name') or '').strip()

            if approver_role not in LEAVE_APPROVER_ROLE_KEYS:
                flash('Please select a valid Leave approver role.', category='error')
                return redirect(url_for('admin.settings'))
            if not user_id:
                flash('Please select a user.', category='error')
                return redirect(url_for('admin.settings'))

            user = User.query.get(user_id)
            if not user:
                flash('User not found.', category='error')
                return redirect(url_for('admin.settings'))
            if user.role == 'admin':
                flash('Admin users cannot be reassigned from approval settings.', category='error')
                return redirect(url_for('admin.settings'))

            if approver_role == 'department':
                if not department_name:
                    flash('Please choose the department for this approver.', category='error')
                    return redirect(url_for('admin.settings'))
                selected_department = Department.query.filter(
                    func.lower(Department.department_name) == department_name.lower()
                ).first()
                if not selected_department:
                    flash('Selected department was not found.', category='error')
                    return redirect(url_for('admin.settings'))
                department_name = selected_department.department_name
            else:
                department_name = None

            if approver_role == 'department':
                existing_role_assignment = LeaveApprover.query.filter(
                    LeaveApprover.user_id != user.id,
                    LeaveApprover.approver_role == 'department',
                    func.lower(LeaveApprover.department_name) == department_name.lower(),
                ).first()
                if existing_role_assignment:
                    flash('That department already has a Department Leave approver.', category='error')
                    return redirect(url_for('admin.settings'))

            existing_assignment = LeaveApprover.query.filter_by(user_id=user.id).first()
            if existing_assignment:
                existing_assignment.approver_role = approver_role
                existing_assignment.department_name = department_name
            else:
                db.session.add(
                    LeaveApprover(
                        user_id=user.id,
                        approver_role=approver_role,
                        department_name=department_name,
                    )
                )

            role_label = dict(LEAVE_APPROVER_ROLES)[approver_role]
            try:
                db.session.commit()
                flash(f'{user.username} assigned as {role_label}.', category='success')
            except Exception:
                db.session.rollback()
                flash('Failed to assign Leave approver.', category='error')
            return redirect(url_for('admin.settings'))

        elif action == 'remove_leave_approver':
            user_id = request.form.get('user_id')
            user = User.query.get(user_id)
            if not user:
                flash('User not found.', category='error')
                return redirect(url_for('admin.settings'))
            assignment = LeaveApprover.query.filter_by(user_id=user.id).first()
            if not assignment:
                flash('Selected user is not a Leave approver.', category='error')
                return redirect(url_for('admin.settings'))

            previous_role = dict(LEAVE_APPROVER_ROLES).get(
                assignment.approver_role,
                'Leave',
            )
            db.session.delete(assignment)
            try:
                db.session.commit()
                flash(f'{user.username} removed from {previous_role} approvers.', category='success')
            except Exception:
                db.session.rollback()
                flash('Failed to remove Leave approver.', category='error')
            return redirect(url_for('admin.settings'))

        elif action == 'add_approver':
            user_id = request.form.get('user_id')
            can_discount = 'can_approve_discount' in request.form
            can_charge = 'can_approve_charge' in request.form

            if not user_id:
                flash('Please select a user.', category='error')
                return redirect(url_for('admin.settings'))

            user = User.query.get(user_id)
            if not user:
                flash('User not found.', category='error')
                return redirect(url_for('admin.settings'))

            existing = PerkApprover.query.filter_by(user_id=user.id).first()
            if existing:
                flash(f'{user.username} is already a perk approver.', category='error')
                return redirect(url_for('admin.settings'))

            new_approver = PerkApprover(
                user_id=user.id,
                can_approve_discount=can_discount,
                can_approve_charge=can_charge,
            )
            db.session.add(new_approver)
            try:
                db.session.commit()
                flash(f'{user.username} added as perk approver.', category='success')
            except Exception:
                db.session.rollback()
                flash('Failed to add approver.', category='error')
            return redirect(url_for('admin.settings'))

        elif action == 'remove_approver':
            approver_id = request.form.get('approver_id')
            approver = PerkApprover.query.get(approver_id)
            if approver:
                db.session.delete(approver)
                try:
                    db.session.commit()
                    flash('Approver removed.', category='success')
                except Exception:
                    db.session.rollback()
                    flash('Failed to remove approver.', category='error')
            else:
                flash('Approver not found.', category='error')
            return redirect(url_for('admin.settings'))

        elif action == 'update_approver':
            approver_id = request.form.get('approver_id')
            approver = PerkApprover.query.get(approver_id)
            if approver:
                approver.can_approve_discount = 'can_approve_discount' in request.form
                approver.can_approve_charge = 'can_approve_charge' in request.form
                try:
                    db.session.commit()
                    flash('Approver permissions updated.', category='success')
                except Exception:
                    db.session.rollback()
                    flash('Failed to update approver.', category='error')
            else:
                flash('Approver not found.', category='error')
            return redirect(url_for('admin.settings'))

    # GET
    approvers = PerkApprover.query.order_by(PerkApprover.id.desc()).all()
    approver_user_ids = [a.user_id for a in approvers]
    esarf_approvers_by_role = {
        role: (
            EsarfApprover.query
            .join(EsarfApprover.user)
            .filter(EsarfApprover.approver_role == role)
            .order_by(User.username.asc())
            .all()
        )
        for role, _label in ESARF_APPROVER_ROLES
    }
    esarf_approver_user_ids = [
        user_id for user_id, in db.session.query(EsarfApprover.user_id).all()
    ]
    esarf_eligible_users = User.query.filter(
        User.role != 'admin',
        User.id.notin_(esarf_approver_user_ids),
    ).order_by(User.username.asc()).all()
    departments = Department.query.order_by(Department.department_name.asc()).all()
    leave_approvers_by_role = {
        role: (
            LeaveApprover.query
            .join(LeaveApprover.user)
            .filter(LeaveApprover.approver_role == role)
            .order_by(User.username.asc())
            .all()
        )
        for role, _label in LEAVE_APPROVER_ROLES
    }
    leave_approver_user_ids = [
        user_id for user_id, in db.session.query(LeaveApprover.user_id).all()
    ]
    leave_eligible_users = User.query.filter(
        User.role != 'admin',
        User.id.notin_(leave_approver_user_ids),
    ).order_by(User.username.asc()).all()
    # Users eligible to be approvers (non-admin users that are not already approvers)
    eligible_users = User.query.filter(
        User.id.notin_(approver_user_ids),
        User.role != 'admin',
    ).order_by(User.username.asc()).all()

    return render_template(
        'admin/settings.html',
        approvers=approvers,
        eligible_users=eligible_users,
        esarf_approver_roles=ESARF_APPROVER_ROLES,
        esarf_approvers_by_role=esarf_approvers_by_role,
        esarf_eligible_users=esarf_eligible_users,
        leave_approver_roles=LEAVE_APPROVER_ROLES,
        leave_approvers_by_role=leave_approvers_by_role,
        leave_eligible_users=leave_eligible_users,
        departments=departments,
    )
